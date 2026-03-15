import os, glob, warnings
import numpy as np
import pandas as pd
from scipy.integrate import cumulative_trapezoid
from scipy.stats import linregress, zscore
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec

warnings.filterwarnings("ignore")

DATA_DIR   = "raw/data"
REPORT_DIR = "reports"
CHART_DIR  = os.path.join(REPORT_DIR, "charts")
REPORT_OUT = os.path.join(REPORT_DIR, "vibration_report.xlsx")

ISO_LIMITS = {
    "Group1_Rigid": [2.3, 4.5, 7.1],   # >75 kW, rigid mount
    "Group2_Rigid": [1.4, 2.8, 4.5],
}
MACHINE_GROUP = "Group1_Rigid"
LIMITS = ISO_LIMITS[MACHINE_GROUP]

MONTH_ORDER = {m: i for i, m in enumerate([
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
])}

os.makedirs(CHART_DIR, exist_ok=True)


# ─────────────────────────────────────────────
#  1. PARSER  —  อ่านไฟล์ .txt → DataFrame
# ─────────────────────────────────────────────
def parse_waveform_file(path):
    """แปลงไฟล์ .txt format 4 คอลัมน์ → (time_ms, accel_g)"""
    try:
        df_raw = pd.read_csv(path, skiprows=15, sep=r'\s+',
                             header=None, on_bad_lines="skip")
        df_raw = df_raw.apply(pd.to_numeric, errors="coerce")

        parts = []
        for col_t, col_a in [(0, 1), (2, 3)]:
            if col_t < len(df_raw.columns) and col_a < len(df_raw.columns):
                part = df_raw[[col_t, col_a]].copy()
                part.columns = ["time_ms", "accel_g"]
                parts.append(part)

        if not parts:
            return None

        df = pd.concat(parts, ignore_index=True).dropna()
        df = df[df["time_ms"].between(-1e6, 1e6)]
        df = df[df["accel_g"].between(-100, 100)]
        df = df.sort_values("time_ms").reset_index(drop=True)
        return df
    except Exception as e:
        print(f"  [WARN] parse error {os.path.basename(path)}: {e}")
        return None


# ─────────────────────────────────────────────
#  2. FEATURE ENGINEERING
# ─────────────────────────────────────────────
def extract_features(df):
    """คำนวณ feature จาก time-domain waveform"""
    t = df["time_ms"].values / 1000.0
    a_raw = df["accel_g"].values
    a = (a_raw - np.mean(a_raw)) * 9.81          # remove DC, convert to m/s²

    # Integrate to velocity (mm/s)
    v = cumulative_trapezoid(a, t, initial=0) * 1000.0

    eps = 1e-9
    rms       = float(np.sqrt(np.mean(v**2)))
    peak      = float(np.max(np.abs(v)))
    crest     = peak / (rms + eps)
    kurt      = float(pd.Series(v).kurt())
    skew      = float(pd.Series(v).skew())
    peak_accel= float(np.max(np.abs(a_raw)))

    return {
        "rms_velocity_mms":  round(rms,   3),
        "peak_velocity_mms": round(peak,  3),
        "crest_factor":      round(crest, 3),
        "kurtosis":          round(kurt,  3),
        "skewness":          round(skew,  3),
        "peak_accel_g":      round(peak_accel, 3),
        "n_samples":         len(df),
    }


# ─────────────────────────────────────────────
#  3. ISO CLASSIFICATION
# ─────────────────────────────────────────────
def iso_classify(rms):
    labels = ["A: Good", "B: Satisfactory", "C: Unsatisfactory", "D: Unacceptable"]
    colors = ["#27ae60",  "#f39c12",          "#e67e22",            "#e74c3c"]
    for i, lim in enumerate(LIMITS):
        if rms <= lim:
            return labels[i], colors[i]
    return labels[-1], colors[-1]


# ─────────────────────────────────────────────
#  4. TREND ANALYSIS
# ─────────────────────────────────────────────
def analyze_trend(monthly_rms):
    """
    monthly_rms: list of (month_idx, rms) tuples, >=2 points
    Returns slope, r², projected months to cross each threshold
    """
    if len(monthly_rms) < 2:
        return {"slope": 0, "r2": 0, "trend_label": "ข้อมูลน้อยเกินไป",
                "months_to_C": None, "months_to_D": None}

    x = np.array([m[0] for m in monthly_rms], dtype=float)
    y = np.array([m[1] for m in monthly_rms], dtype=float)
    slope, intercept, r, _, _ = linregress(x, y)
    r2 = r**2

    def months_to_threshold(thresh):
        if slope <= 0:
            return None
        val = (thresh - intercept) / slope
        remaining = val - x[-1]
        return round(remaining, 1) if remaining > 0 else 0

    if slope > 0.15:
        trend_label = "⬆ แนวโน้มเพิ่มขึ้น (เฝ้าระวัง)"
    elif slope > 0.05:
        trend_label = "↗ เพิ่มขึ้นเล็กน้อย"
    elif slope < -0.05:
        trend_label = "↘ ลดลง (ดีขึ้น)"
    else:
        trend_label = "→ คงที่"

    return {
        "slope":         round(float(slope), 4),
        "r2":            round(float(r2), 3),
        "trend_label":   trend_label,
        "months_to_C":   months_to_threshold(LIMITS[2]),   # 7.1
        "months_to_D":   months_to_threshold(LIMITS[2]),   # same, flag earlier
    }


# ─────────────────────────────────────────────
#  5. ANOMALY DETECTION  (Z-score)
# ─────────────────────────────────────────────
def detect_anomalies(df_summary):
    """เพิ่มคอลัมน์ anomaly_flag ด้วย Z-score ต่อเครื่อง"""
    df_summary = df_summary.copy()
    df_summary["anomaly_flag"] = False
    df_summary["zscore_rms"]   = 0.0

    for machine, grp in df_summary.groupby("machine"):
        if len(grp) < 3:
            continue
        z = zscore(grp["rms_velocity_mms"].astype(float))
        idx = grp.index
        df_summary.loc[idx, "zscore_rms"]   = z.round(2)
        df_summary.loc[idx, "anomaly_flag"] = np.abs(z) > 2.0

    return df_summary


# ─────────────────────────────────────────────
#  6. CHART  —  กราฟ trend ต่อเครื่อง
# ─────────────────────────────────────────────
def plot_machine_trend(machine_name, df_machine):
    """วาดกราฟ RMS trend + ISO zones + anomaly markers"""
    df = df_machine.sort_values("month_idx").copy()
    if df.empty:
        return None

    fig = plt.figure(figsize=(12, 5), facecolor="#0f1117")
    ax = fig.add_subplot(111)
    ax.set_facecolor("#0f1117")

    x    = df["month_idx"].values
    y    = df["rms_velocity_mms"].values
    labs = df["month"].values

    # ISO zones (shading)
    ax.axhspan(0,        LIMITS[0], alpha=0.12, color="#27ae60", zorder=1)
    ax.axhspan(LIMITS[0],LIMITS[1], alpha=0.12, color="#f39c12", zorder=1)
    ax.axhspan(LIMITS[1],LIMITS[2], alpha=0.12, color="#e67e22", zorder=1)
    ax.axhspan(LIMITS[2], max(y.max()*1.3, LIMITS[2]+2),
               alpha=0.12, color="#e74c3c", zorder=1)

    # Zone lines
    for lim, col, lbl in zip(LIMITS,
                              ["#27ae60","#f39c12","#e74c3c"],
                              ["A/B  2.3","B/C  4.5","C/D  7.1"]):
        ax.axhline(lim, color=col, lw=0.8, linestyle="--", alpha=0.6)
        ax.text(x[-1]+0.1, lim+0.05, lbl, color=col,
                fontsize=7.5, va="bottom", alpha=0.8)

    # Trend line
    if len(x) >= 2:
        z = np.polyfit(x, y, 1)
        p = np.poly1d(z)
        x_ext = np.linspace(x[0], x[-1]+3, 60)
        ax.plot(x_ext, p(x_ext), "--", color="#aaaaaa",
                lw=1, alpha=0.5, label="Trend line")

    # RMS line
    ax.plot(x, y, "-o", color="#4fc3f7", lw=2,
            markersize=5, markerfacecolor="white",
            markeredgecolor="#4fc3f7", zorder=3, label="RMS Velocity")

    # Anomaly markers
    anom = df[df["anomaly_flag"] == True]
    if not anom.empty:
        ax.scatter(anom["month_idx"], anom["rms_velocity_mms"],
                   s=120, color="#ff4757", marker="*",
                   zorder=5, label="Anomaly detected")

    # X labels
    tick_pos = list(range(len(labs)))
    ax.set_xticks(x)
    ax.set_xticklabels(labs, rotation=30, ha="right",
                       fontsize=9, color="#cccccc")

    ax.set_ylabel("RMS Velocity (mm/s)", color="#cccccc", fontsize=10)
    ax.set_title(f"  {machine_name}", color="#ffffff",
                 fontsize=12, fontweight="bold", loc="left", pad=10)
    ax.tick_params(colors="#888888")
    for spine in ax.spines.values():
        spine.set_edgecolor("#333333")

    # Legend
    patches = [
        mpatches.Patch(color="#27ae60", alpha=0.5, label="A: Good (≤2.3)"),
        mpatches.Patch(color="#f39c12", alpha=0.5, label="B: Satisfactory (≤4.5)"),
        mpatches.Patch(color="#e67e22", alpha=0.5, label="C: Unsatisfactory (≤7.1)"),
        mpatches.Patch(color="#e74c3c", alpha=0.5, label="D: Unacceptable (>7.1)"),
    ]
    legend = ax.legend(handles=patches + [
        plt.Line2D([0],[0],color="#4fc3f7",lw=2,label="RMS Velocity"),
    ], loc="upper left", fontsize=7.5,
                       facecolor="#1a1d27", edgecolor="#333333",
                       labelcolor="#cccccc", ncol=2)

    plt.tight_layout()
    safe_name = machine_name.replace("/","_").replace(" ","_")
    chart_path = os.path.join(CHART_DIR, f"{safe_name}_trend.png")
    plt.savefig(chart_path, dpi=150, bbox_inches="tight",
                facecolor="#0f1117")
    plt.close()
    return chart_path


# ─────────────────────────────────────────────
#  7. EXCEL REPORT
# ─────────────────────────────────────────────
def write_excel_report(df_all, trend_summary, alert_list):
    with pd.ExcelWriter(REPORT_OUT, engine="openpyxl") as writer:
        # Sheet 1: Detail
        df_out = df_all[[
            "file_name","machine","month","month_idx",
            "rms_velocity_mms","peak_velocity_mms",
            "crest_factor","kurtosis",
            "iso_condition","zscore_rms","anomaly_flag"
        ]].copy()
        df_out.to_excel(writer, sheet_name="Detail", index=False)

        # Sheet 2: Trend Summary
        ts = pd.DataFrame(trend_summary)
        ts.to_excel(writer, sheet_name="Trend Summary", index=False)

        # Sheet 3: Alerts
        if alert_list:
            al = pd.DataFrame(alert_list)
            al.to_excel(writer, sheet_name="Alerts", index=False)
        else:
            pd.DataFrame([{"info": "ไม่พบรายการที่ต้องแจ้งเตือน"}])\
              .to_excel(writer, sheet_name="Alerts", index=False)

        # Color coding
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = writer.book
        ws = wb["Detail"]

        color_map = {
            "A: Good":           ("C8F7C5","1e8449"),
            "B: Satisfactory":   ("FEF9E7","d68910"),
            "C: Unsatisfactory": ("FDEBD0","ca6f1e"),
            "D: Unacceptable":   ("FADBD8","922b21"),
        }
        for row in ws.iter_rows(min_row=2):
            cond_cell = None
            for cell in row:
                if cell.column == 9:   # iso_condition column
                    cond_cell = cell
            if cond_cell and cond_cell.value in color_map:
                bg, fg = color_map[cond_cell.value]
                fill = PatternFill("solid", fgColor=bg)
                font = Font(color=fg, bold=True)
                for cell in row:
                    cell.fill = fill
                cond_cell.font = font

        # Auto width
        for ws_name in ["Detail","Trend Summary","Alerts"]:
            ws_cur = wb[ws_name]
            for col in ws_cur.columns:
                max_w = max(len(str(c.value or "")) for c in col)
                ws_cur.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)

    print(f"  ✅ บันทึก Excel: {REPORT_OUT}")


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print("\n" + "="*60)
    print("  Vibration Predictive Maintenance Pipeline")
    print("="*60)

    # ── Load all files ──
    files = sorted(glob.glob(os.path.join(DATA_DIR, "*.txt")))
    if not files:
        print(f"\n❌ ไม่พบไฟล์ .txt ใน {DATA_DIR}")
        print("   ลองรัน: python raw/data/gen_sample_data.py  ก่อน")
        return

    print(f"\n📂 พบไฟล์ทั้งหมด: {len(files)} ไฟล์\n")

    records = []
    for path in files:
        fname = os.path.basename(path)
        parts = fname.replace(".txt","").split("__")
        machine = parts[0].lstrip("A_").strip() if parts else fname
        month_raw = parts[1] if len(parts) > 1 else "Unknown"

        # month_idx สำหรับ sort
        month_name = month_raw[:3]
        year_suffix = month_raw[3:] if len(month_raw) > 3 else "24"
        year_num = int("20" + year_suffix) if year_suffix.isdigit() else 2024
        month_idx = year_num * 12 + MONTH_ORDER.get(month_name, 0)

        df = parse_waveform_file(path)
        if df is None or len(df) < 10:
            print(f"  [SKIP] {fname}")
            continue

        feats = extract_features(df)
        cond, _ = iso_classify(feats["rms_velocity_mms"])

        records.append({
            "file_name":    fname,
            "machine":      machine,
            "month":        month_raw,
            "month_idx":    month_idx,
            "iso_condition": cond,
            **feats,
        })
        print(f"  ✓ {fname:<50} RMS={feats['rms_velocity_mms']:>6.3f} mm/s  [{cond}]")

    if not records:
        print("\n❌ ไม่สามารถประมวลผลไฟล์ได้เลย")
        return

    df_all = pd.DataFrame(records).sort_values(["machine","month_idx"])

    # ── Anomaly detection ──
    df_all = detect_anomalies(df_all)

    # ── Trend analysis + Charts ──
    print(f"\n{'─'*60}")
    print("📈  Trend Analysis & Charts")
    print(f"{'─'*60}")

    trend_summary = []
    alert_list    = []

    for machine, grp in df_all.groupby("machine"):
        grp_s = grp.sort_values("month_idx")
        monthly_rms = list(zip(grp_s["month_idx"], grp_s["rms_velocity_mms"]))

        trend = analyze_trend(monthly_rms)
        latest_rms  = float(grp_s["rms_velocity_mms"].iloc[-1])
        latest_cond, _ = iso_classify(latest_rms)

        trend_summary.append({
            "machine":          machine,
            "latest_month":     grp_s["month"].iloc[-1],
            "latest_rms":       latest_rms,
            "latest_condition": latest_cond,
            "trend":            trend["trend_label"],
            "slope_per_month":  trend["slope"],
            "r_squared":        trend["r2"],
            "est_months_to_C":  trend["months_to_C"],
            "anomaly_count":    int(grp_s["anomaly_flag"].sum()),
        })

        # Alerts
        if latest_cond in ["C: Unsatisfactory","D: Unacceptable"]:
            alert_list.append({
                "priority": "🔴 HIGH" if latest_cond == "D: Unacceptable" else "🟠 MEDIUM",
                "machine":  machine,
                "latest_rms": latest_rms,
                "condition":  latest_cond,
                "trend":      trend["trend_label"],
                "months_to_D": trend["months_to_D"],
                "recommendation": "ตรวจสอบทันที" if latest_cond == "D: Unacceptable"
                                  else "วางแผน maintenance ภายใน 1–2 เดือน"
            })

        # Chart
        chart_path = plot_machine_trend(machine, grp_s)
        if chart_path:
            print(f"  📊 {machine:<45} → {os.path.basename(chart_path)}")

    # ── Excel ──
    print(f"\n{'─'*60}")
    print("📝  Generating Excel Report")
    print(f"{'─'*60}")
    write_excel_report(df_all, trend_summary, alert_list)

    # ── Console Summary ──
    print(f"\n{'='*60}")
    print("  SUMMARY")
    print(f"{'='*60}")
    print(f"  เครื่องที่วิเคราะห์:   {df_all['machine'].nunique()} เครื่อง")
    print(f"  ไฟล์ทั้งหมด:           {len(df_all)} ไฟล์")
    print(f"  พบ Anomaly:            {int(df_all['anomaly_flag'].sum())} ครั้ง\n")

    print(f"  {'เครื่องจักร':<45} {'RMS':>8}  สถานะ")
    print(f"  {'─'*45} {'─'*8}  {'─'*25}")
    for t in sorted(trend_summary, key=lambda x: -x["latest_rms"]):
        icon = "🔴" if "Unacceptable" in t["latest_condition"] \
               else "🟠" if "Unsatisfactory" in t["latest_condition"] \
               else "🟡" if "Satisfactory"   in t["latest_condition"] \
               else "🟢"
        print(f"  {t['machine']:<45} {t['latest_rms']:>6.2f}  {icon} {t['latest_condition']}")

    if alert_list:
        print(f"\n  ⚠️  Alerts ({len(alert_list)} รายการ):")
        for a in alert_list:
            print(f"     {a['priority']}  {a['machine']}")
            print(f"              {a['recommendation']}")

    print(f"\n  📁 Output:")
    print(f"     {REPORT_OUT}")
    print(f"     {CHART_DIR}/*.png")
    print(f"\n{'='*60}\n")


if __name__ == "__main__":
    main()
